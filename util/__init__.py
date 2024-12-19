"""General utility functions module.

This module contains a collection of general utility functions that are employed
across various projects to streamline and enhance coding practices. These functions
are designed to be reusable, efficient, and compatible with the common requirements
of different projects within the organization.
"""

import os
import time
import json
import smtplib
import warnings
from enum import Enum
from pathlib import Path
from email.utils import COMMASPACE
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border
import win32com.client as win32
import markdown


# pylint: disable=R0913


smtp_host: str | None = None
smtp_port: int | None = None
smtp_username: str | None = None
smtp_password: str | None = None
smtp_api_key: str | None = None
smtp_from: str = 'alan.baker@imarcgroup.info'


class EmailMode(Enum):
    """Enum for email modes."""
    OUTLOOK="outlook"
    API="api"


class TemplateFormat(Enum):
    MD="md"
    PLAIN="plain"


class TemplateOutputFormat(Enum):
    PLAIN="plain"
    HTML="html"


def _outlook_mailing(
        subject: str,
        message: str,
        recipients: list[str],
        mail_type: str = "plain",
        attachments: list[str] = None,
        cc: list[str] = None,
        bcc: list[str] = None,
):

    outlook = win32.Dispatch('outlook.application')

    # for email in recipients:
    mail = outlook.CreateItem(0)

    # mail.To = str(email)
    mail.To = ";".join(recipients)
    if cc:
        mail.CC = ";".join(cc)

    if bcc:
        mail.BCC = ";".join(bcc)

    mail.Subject = str(subject)

    if attachments:
        mail.attachments = attachments

    if mail_type == "plain":
        mail.Body = str(message)
    elif mail_type == "html":
        mail.HTMLBody = str(message)
    else:
        raise ValueError(
            "Invalid mail type. Choose 'plain' or 'html'.")

    mail.Send()


def _api_mailing(
        subject: str,
        message: str,
        recipients: list[str],
        mail_type: str = "plain",
        attachments: list[str] = None,
        cc: list[str] = None,
        bcc: list[str] = None,
):
    if (
        smtp_api_key is None
        and (smtp_username is None
                and smtp_password is None)
    ):
        raise ValueError(
            "Please provide either an smtp_api_key or "
            "(smtp_username and smtp_password)")

    mail_from = smtp_from

    # for email in recipients:
    # mail_to = str(email)

    msg = MIMEMultipart()

    msg['From'] = mail_from
    msg['To'] = COMMASPACE.join(recipients)

    if cc:
        msg['Cc'] = COMMASPACE.join(cc)

    if bcc:
        msg['Bcc'] = COMMASPACE.join(bcc)

    msg['Subject'] = str(subject)

    html_part = MIMEText(str(message), mail_type)
    msg.attach(html_part)

    for file_path in attachments or []:
        with open(file_path, "rb") as file:
            part = MIMEApplication(
                file.read(),
                Name=os.path.basename(file_path)
            )
        # After the file is closed
        part['Content-Disposition'] = 'attachment; '\
            f'filename="{os.path.basename(file_path)}"'
        msg.attach(part)

    server = smtplib.SMTP_SSL(smtp_host, smtp_port)
    server.ehlo()

    if smtp_api_key:
        server.login('apikey', smtp_api_key)
    else:
        server.login(smtp_username, smtp_password)

    server.sendmail(mail_from, recipients, msg.as_string())
    server.close()

    time.sleep(1)


def send_mail(
        subject: str,
        message: str,
        recipients: list[str],
        mail_type: str = "plain",
        attachments: list[str] = None,
        mode: EmailMode | list[EmailMode] = None,
        cc: list[str] = None,
        bcc: list[str] = None,
):
    """sends mail

    Args:
        mail_type:
        recipients:
        subject (str): subject to send
        message (str): message to send
        recipients (string): email addresses to send to
    """
    if not mode:
        mode = [EmailMode.OUTLOOK, EmailMode.API]

    if mode == EmailMode.OUTLOOK:
        _outlook_mailing(
            subject = subject,
            message = message,
            recipients = recipients,
            mail_type = mail_type,
            attachments = attachments,
            cc = cc,
            bcc = bcc,
        )

    elif mode == EmailMode.API:
        _api_mailing(
            subject = subject,
            message = message,
            recipients = recipients,
            mail_type = mail_type,
            attachments = attachments,
            cc = cc,
            bcc = bcc,
        )

    elif isinstance(mode, list):
        for mode_ in mode:
            try:
                send_mail(
                    subject = subject,
                    message = message,
                    recipients = recipients,
                    mail_type = mail_type,
                    attachments = attachments,
                    mode = mode_,
                    cc = cc,
                    bcc = bcc,
                )
                break

            # pylint: disable-next=W0718
            except Exception:
                warnings.warn("Failed to send mail via '{mode_}' mode.")

    else:
        raise ValueError("Invalid email mode. Choose 'outlook' or 'api'.")


def touch_excel(
    df: pd.DataFrame,
    file_path: str | Path,
    sheet_name: str = "Sheet1",
    add_df: pd.DataFrame = None,
):
    """this function updates or creates a new sheet with the given dataframe

    if the file does not exist yet, it will be created

    It can also concatenate two different DataFrames and merge them into one
    them write them in the given sheet

    Args:
        df (pd.DataFrame): 
            dataframe that is to be written to the sheet

        file_path (string): 
            path to the file to be written

        sheet_name (string, optional): 
            name of the sheet that is to updated. Defaults to "Sheet1".

        add_df (pd.DataFrame, optional): 
            other dataframe that is to be merged. Defaults to None.

    Raises:
        PermissionError: 
            When the file is opened by user and is currently being used. 
            Please Close the file
    """
    if isinstance(file_path, Path):
        file_path = str(file_path)

    if add_df is not None:
        df = pd.concat([df, add_df], ignore_index=True)

    try:
        if not os.path.exists(file_path):
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(
                file_path,
                mode="a",
                engine='openpyxl',
                if_sheet_exists='replace',
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                time.sleep(5)
                # writer.close()
    except PermissionError as e:
        e.message = "File might be open. Close it."
        raise e


def get_config(
    file_name: str,
    is_global: bool = False,
    file_type: str = "json",
    encoding: str = "utf-8",
) -> dict:
    """
    Load configuration from a specified JSON or YAML file.

    This function reads the configuration from a file (either JSON or YAML)
    located either in the current working directory or in a global 
    "Project Configurations" directory in the user's home folder, depending 
    on the `is_global` flag.

    Args:
        file_name (str): The name of the file from which to load configurations.
        is_global (bool, optional): Flag to determine the location from which
            to load the configfurations. If True, the configuration is loaded
            from the user's home directory. Otherwise, it's loaded from the
            current working directory. Defaults to False.
        file_type (str, optional): The type of the file to read from. Can be
            either 'json' or 'yaml'. Defaults to 'json'.
        encoding (str, optional): The encoding to use when opening the file.
            Defaults to 'utf-8'.

    Returns:
        dict: A dictionary containing the loaded configurations.

    Raises:
        FileNotFoundError: If the specified file does not exist in the expected
            location.
        ValueError: If the file_type is neither 'json' nor 'yaml'.
    """
    if is_global:
        base_path = os.path.join(
            os.path.expanduser('~'), "Project Configurations")
    else:
        base_path = os.getcwd()

    config_path = os.path.join(
        base_path,
        file_name,
    )

    if not os.path.exists(config_path):
        raise FileNotFoundError(f"{file_name} Not found at {config_path}")

    with open(config_path, 'r', encoding=encoding) as f:
        if file_type == "json":
            config = json.load(f)
        elif file_type == "yaml":
            config = yaml.load(f.read(), Loader=yaml.FullLoader)
        else:
            raise ValueError(f"File type {file_type} not supported")

    return config


def style_excel(
    path: str,
    sheet_name: str | list[str] = None,
    header_color: str = 'D0EFFF',
    bold: bool = True,
    font_size: int = None,
):
    """
    Applies a header color, font size, and bold style to the first row of an Excel file.

    Args:
        path (str): The path to the Excel file.
        header_color (str, optional): The color to use for the header. Defaults to 'D0EFFF'.
        bold (bool, optional): Whether to apply bold style. Defaults to True.
        font_size (int, optional): The font size to use. Defaults to 12.
    """
    workbook = load_workbook(path)

    sheets = sheet_name

    if not sheet_name:
        sheets = [workbook.active.title]

    if isinstance(sheet_name, str):
        sheets = [sheet_name]

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in sheets:

        if sheet not in workbook.sheetnames:
            raise ValueError(f"Could not find sheet '{sheet}' in workbook")

        input_worksheet = workbook[sheet]

        for column in input_worksheet.iter_cols():

            col_width = 5

            column_letter = column[0].column_letter

            column[0].fill = PatternFill(
                fill_type='solid',
                start_color=header_color,
                end_color=header_color,
            )

            font_style = {}

            if bold:
                font_style['bold'] = True

            if font_size:
                font_style['size'] = font_size

            column[0].font = Font(**font_style)

            column[0].border = thin_border

            if len(str(column[0].value)) > col_width:
                col_width = len(column[0].value)

            adjusted_width = (col_width + 2) * 1.2

            input_worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(path)

def fill_template(
    path: str,
    output_format: TemplateOutputFormat | str = TemplateOutputFormat.HTML,
    verbose: bool = False,
    *args,
    **kwargs,
):

    if isinstance(output_format, str):
        output_format = TemplateOutputFormat(output_format)
    assert isinstance(output_format, TemplateOutputFormat), "Invalid output format provided."

    if verbose: print(path)

    if not os.path.exists(path):
        raise FileNotFoundError("No template file present.")

    with open(path, 'r', encoding='utf-8') as f:
        template = f.read()
        output_template = template.format(*args, **kwargs)
        if output_format == TemplateOutputFormat.HTML:
            output_template = markdown.markdown(output_template)

        return output_template
